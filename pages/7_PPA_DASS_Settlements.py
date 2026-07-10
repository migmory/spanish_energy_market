# pages/7_PPA_DASS_Settlements.py
"""
PPA & DASS Settlements — offtaker-facing settlement dashboard.

Data lineage (single source of truth = data/PV___BESS_settlement_app.xlsx):
- When the Aurora/workbook curve is selected, monthly and annual captured solar prices,
  BESS revenues and TB4 spreads are read directly from the workbook summary tables.
- Only when the user uploads an hourly forward curve does the app recompute solar capture
  and BESS dispatch from the uploaded prices.
- Performance: the hourly xlsx is converted once to a parquet cache next to it; subsequent
  loads are near-instant.
"""

from __future__ import annotations

import os
import smtplib
import urllib.parse
from email.message import EmailMessage
from io import BytesIO
from pathlib import Path

import altair as alt
import numpy as np
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Page setup & corporate styling
# ---------------------------------------------------------------------------
st.set_page_config(page_title="PPA & DASS Settlements", layout="wide")

GREEN = "#1f8a5f"
GREEN_DARK = "#0f6b47"
GREEN_SOFT = "#e6f4ee"
DASS_GREEN = "#198754"
DASS_GREEN_SOFT = "#e3f6ec"
ORANGE = "#e8862e"
RED = "#cf4d4d"
INK = "#12332a"
MUTED = "#6b7f78"
RTE = 0.925 * 0.925  # ~0.856

st.markdown(
    f"""
    <style>
    .block-container {{ padding-top: 1.1rem; max-width: 1850px; padding-left: 2.2rem; padding-right: 2.2rem; }}
    html, body, [class*="css"] {{ font-family: "Inter","Segoe UI",system-ui,sans-serif; }}

    .nx-hero {{
        background: linear-gradient(120deg, {GREEN_DARK} 0%, {GREEN} 55%, #2fae79 100%);
        border-radius: 18px; padding: 26px 32px; color: #fff; margin-bottom: 18px;
        box-shadow: 0 8px 24px rgba(15,107,71,.18);
    }}
    .nx-hero h1 {{ font-size: 2.0rem; font-weight: 800; margin: 0; letter-spacing:-.02em; color:#fff; }}
    .nx-hero p  {{ margin: 6px 0 0 0; color: #d9efe6; font-size: .95rem; max-width: 940px; }}
    .nx-pill {{
        display:inline-block; background:#fff; color:{GREEN_DARK}; font-weight:700;
        font-size:.75rem; padding:4px 12px; border-radius:999px; margin-bottom:10px;
        letter-spacing:.06em; text-transform:uppercase;
    }}

    .nx-section {{ display:flex; align-items:center; gap:10px; margin: 28px 0 4px 0; }}
    .nx-dot {{ width:13px; height:13px; border-radius:4px; display:inline-block; }}
    .nx-section h2 {{ font-size:1.45rem; font-weight:800; color:{INK}; margin:0; letter-spacing:-.01em; }}
    .nx-sub {{ color:{MUTED}; font-size:.9rem; margin: 0 0 12px 23px; }}

    .kpi {{
        background:#fff; border:1px solid #e5eeea; border-radius:18px;
        padding:22px 24px 20px 24px; box-shadow: 0 3px 12px rgba(18,51,42,.06); height:100%;
        text-align:center; display:flex; flex-direction:column; justify-content:center; min-height:138px;
    }}
    .kpi .label {{ color:{MUTED}; font-size:.8rem; font-weight:700; text-transform:uppercase; letter-spacing:.07em; }}
    .kpi .value {{ font-size:3.1rem; font-weight:800; letter-spacing:-.035em; line-height:1.05; margin-top:4px; }}
    .kpi .unit  {{ font-size:1.15rem; font-weight:700; color:{MUTED}; margin-left:5px; }}
    .kpi .foot  {{ color:{MUTED}; font-size:.8rem; margin-top:8px; }}
    .pos {{ color:{GREEN_DARK}; }} .neg {{ color:{RED}; }} .neu {{ color:{INK}; }} .pur {{ color:{DASS_GREEN}; }}

    .offtaker-badge {{
        display:inline-block; background:{GREEN_SOFT}; color:{GREEN_DARK};
        border:1px solid #cfe8dc; font-weight:700; font-size:.8rem;
        padding:5px 14px; border-radius:999px; letter-spacing:.03em;
    }}
    .offtaker-badge.dass {{ background:{DASS_GREEN_SOFT}; color:{DASS_GREEN}; border-color:#c8ecd7; }}

    /* segmented-control look for horizontal radios */
    div[role="radiogroup"] {{ gap: 6px; }}
    div[role="radiogroup"] > label {{
        background:#f4f8f6; border:1px solid #dbe7e1; border-radius:999px;
        padding:4px 14px 4px 8px; transition: all .15s ease;
    }}
    div[role="radiogroup"] > label:has(input:checked) {{
        background:{GREEN_SOFT}; border-color:{GREEN}; font-weight:700;
    }}
    /* slider accent */
    div[data-baseweb="slider"] div[role="slider"] {{ background:{GREEN_DARK}; }}
    div[data-testid="stExpander"] {{ border-radius: 14px; border:1px solid #e5eeea; }}


    .nx-module-banner {{
        margin: 26px 0 14px 0;
        padding: 18px 22px;
        border-radius: 18px;
        border: 1px solid #cfe8dc;
        background: linear-gradient(120deg, #f7fcfa 0%, #e6f4ee 100%);
        box-shadow: 0 6px 18px rgba(18,51,42,.07);
        display: flex;
        align-items: center;
        justify-content: space-between;
        gap: 18px;
    }}
    .nx-module-banner.dass {{
        background: linear-gradient(120deg, #f7fcfa 0%, #e3f6ec 100%);
        border-color: #c8ecd7;
    }}
    .nx-module-title {{ font-size: 1.28rem; font-weight: 900; color: #12332a; letter-spacing: -.01em; }}
    .nx-module-subtitle {{ margin-top: 3px; color: #6b7f78; font-size: .9rem; }}
    .nx-module-tag {{
        padding: 8px 14px;
        border-radius: 999px;
        border: 1.5px solid #0f6b47;
        color: #0f6b47;
        font-size: .76rem;
        font-weight: 900;
        letter-spacing: .07em;
        white-space: nowrap;
        text-transform: uppercase;
    }}
    .nx-chart-title {{
        margin: 20px 0 6px 0;
        padding-left: 10px;
        border-left: 5px solid #0f6b47;
        color: #12332a;
        font-size: 1.02rem;
        font-weight: 900;
    }}
    .nx-chart-title.dass {{ border-left-color: #198754; }}
    .nx-chart-note {{ color:#6b7f78; font-size:.82rem; margin: -2px 0 8px 15px; }}
    .nx-print-card {{
        margin-top: 34px;
        padding: 18px 22px;
        border: 1px solid #cfe8dc;
        border-radius: 18px;
        background: linear-gradient(120deg, #f7fcfa 0%, #e6f4ee 100%);
        display: flex;
        justify-content: space-between;
        align-items: center;
        gap: 18px;
        box-shadow: 0 3px 14px rgba(18,51,42,.06);
    }}
    .nx-print-title {{ font-size: 1.25rem; font-weight: 850; color: #12332a; }}
    .nx-print-subtitle {{ font-size: .9rem; color: #6b7f78; margin-top: 4px; }}
    .nx-print-seal {{
        border: 2px solid #0f6b47;
        color: #0f6b47;
        border-radius: 999px;
        padding: 10px 18px;
        font-size: .82rem;
        font-weight: 900;
        letter-spacing: .08em;
        white-space: nowrap;
    }}


    .nx-module-wrap {{
        padding: 18px 20px 22px 20px;
        border-radius: 24px;
        margin: 28px 0 34px 0;
        box-shadow: 0 10px 24px rgba(18,51,42,.06);
    }}
    .nx-module-wrap.solar {{
        background: linear-gradient(180deg, #fff8df 0%, #fffdf5 100%);
        border: 1px solid #efd98f;
    }}
    .nx-module-wrap.dass {{
        background: linear-gradient(180deg, #eef9f2 0%, #fbfffc 100%);
        border: 1px solid #bfe0cb;
    }}
    .nx-section-card-note {{
        margin: 0 0 8px 0;
        padding: 10px 14px;
        border-radius: 12px;
        font-size: .87rem;
        font-weight: 650;
        color: #7d6415;
        background: rgba(255, 238, 174, .45);
        border: 1px solid #edd489;
    }}
    .nx-section-card-note.dass {{
        color: #16653f;
        background: rgba(183, 235, 199, .38);
        border-color: #bfe0cb;
    }}



    .nx-price-grid {{
        display: grid;
        grid-template-columns: repeat(3, minmax(0, 1fr));
        gap: 14px;
        margin: 14px 0 10px 0;
    }}
    .nx-price-card {{
        background: rgba(255,255,255,.92);
        border: 1px solid rgba(15,107,71,.16);
        border-radius: 18px;
        padding: 18px 20px;
        box-shadow: 0 5px 16px rgba(18,51,42,.055);
    }}
    .nx-price-card.solar {{
        border-color: #ead28a;
        background: linear-gradient(180deg, #fffdf5 0%, #fff7db 100%);
    }}
    .nx-price-card.dass {{
        border-color: #bfe0cb;
        background: linear-gradient(180deg, #fbfffc 0%, #e9f8ee 100%);
    }}
    .nx-price-label {{
        font-size: .76rem;
        color: #6b7f78;
        text-transform: uppercase;
        letter-spacing: .08em;
        font-weight: 850;
    }}
    .nx-price-value {{
        margin-top: 5px;
        font-size: 2.75rem;
        line-height: .98;
        font-weight: 950;
        letter-spacing: -.04em;
        color: #12332a;
    }}
    .nx-price-value .unit {{
        font-size: 1.0rem;
        letter-spacing: 0;
        color: #6b7f78;
        margin-left: 4px;
        font-weight: 850;
    }}
    .nx-price-foot {{
        margin-top: 8px;
        color: #6b7f78;
        font-size: .83rem;
        line-height: 1.3;
    }}
    .nx-stepbar {{
        display: grid;
        grid-template-columns: repeat(3, minmax(0, 1fr));
        gap: 10px;
        margin: 10px 0 16px 0;
    }}
    .nx-step {{
        border: 1px solid #e5eeea;
        border-radius: 14px;
        background: rgba(255,255,255,.72);
        padding: 10px 12px;
        color: #12332a;
        font-size: .86rem;
        font-weight: 700;
    }}
    .nx-step span {{
        display:inline-flex;
        align-items:center;
        justify-content:center;
        width:22px;
        height:22px;
        border-radius:999px;
        background:#0f6b47;
        color:#fff;
        font-size:.72rem;
        margin-right:7px;
    }}
    @media (max-width: 900px) {{
        .nx-price-grid, .nx-stepbar {{ grid-template-columns: 1fr; }}
    }}

    @media print {{
        @page {{ size: A4 landscape; margin: 8mm; }}
        .stApp {{ background: white !important; }}
        header, footer, [data-testid="stToolbar"], [data-testid="stSidebar"],
        [data-testid="stDecoration"], .stDeployButton {{ display: none !important; }}
        .block-container {{ max-width: 100% !important; padding: 0.3rem 0.5rem !important; }}
        .nx-hero, .nx-module-banner, .kpi, .nx-print-card {{ break-inside: avoid; box-shadow: none !important; }}
        button {{ display: none !important; }}
        .nx-print-card::after {{
            content: "NEXWELL POWER";
            position: fixed;
            right: 12mm;
            bottom: 8mm;
            color: rgba(15,107,71,.35);
            border: 1.5px solid rgba(15,107,71,.35);
            border-radius: 999px;
            padding: 6px 14px;
            font-size: 10px;
            font-weight: 900;
            letter-spacing: .1em;
        }}
    }}

    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown(
    """
    <div class="nx-hero">
      <span class="nx-pill">Offtaker view · settlements payable to the buyer · FINAL v2024-2030 PDF · FAST CORPORATE UI</span>
      <h1>🛡️ PPA &amp; DASS Settlements</h1>
      <p>What the hedge would have paid you — and what it is expected to pay.
      Historical settlements 2021 – Jun 2026 on OMIE outturn, forward 2027 – 2040 on the
      Aurora nominal curve or your own price scenario. Monthly or annual, in €/MWh · k€/MW and in €.</p>
    </div>
    """,
    unsafe_allow_html=True,
)

# ---------------------------------------------------------------------------
# Constants & paths
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = BASE_DIR / "data"
DATA_FILE_CANDIDATES = [
    DATA_DIR / "PV___BESS_settlement_app.xlsx",
    DATA_DIR / "PV_BESS_settlement_app.xlsx",
    BASE_DIR / "PV___BESS_settlement_app.xlsx",
]

BESS_POWER_MW = 1.0
BESS_CAPACITY_MWH = 4.0
ETA_CH = 0.925
ETA_DIS = 0.925
E_CHARGE_GRID = BESS_CAPACITY_MWH / ETA_CH   # 4.3243 MWh bought / cycle
E_DISCHARGE = BESS_CAPACITY_MWH * ETA_DIS    # 3.7000 MWh sold / cycle
MONTH_ABBR = ["J", "F", "M", "A", "M2", "J2", "Jl", "Au", "S", "O", "N", "D"]
MONTH_LBL = {i + 1: l for i, l in enumerate(
    ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"])}


def resolve_data_file() -> Path | None:
    for p in DATA_FILE_CANDIDATES:
        if p.exists():
            return p
    return None


def kpi(col, label: str, value: str, unit: str = "", foot: str = "", tone: str = "neu"):
    col.markdown(
        f"""<div class="kpi"><div class="label">{label}</div>
        <div class="value {tone}">{value}<span class="unit">{unit}</span></div>
        <div class="foot">{foot}</div></div>""",
        unsafe_allow_html=True,
    )


def section(title: str, color: str, subtitle: str, badge: str, badge_cls: str = ""):
    st.markdown(
        f"""<div class="nx-section"><span class="nx-dot" style="background:{color};"></span>
        <h2>{title}</h2><span class="offtaker-badge {badge_cls}">{badge}</span></div>
        <p class="nx-sub">{subtitle}</p>""",
        unsafe_allow_html=True,
    )


def module_banner(title: str, subtitle: str, tag: str, kind: str = ""):
    cls = "nx-module-banner dass" if kind == "dass" else "nx-module-banner"
    st.markdown(
        f"""<div class="{cls}">
        <div><div class="nx-module-title">{title}</div>
        <div class="nx-module-subtitle">{subtitle}</div></div>
        <div class="nx-module-tag">{tag}</div></div>""",
        unsafe_allow_html=True,
    )


def start_module_wrap(kind: str = "solar"):
    cls = "nx-module-wrap dass" if kind == "dass" else "nx-module-wrap solar"
    st.markdown(f"<div class='{cls}'>", unsafe_allow_html=True)


def end_module_wrap():
    st.markdown("</div>", unsafe_allow_html=True)


def price_card(label: str, value: str, unit: str, foot: str, kind: str = "solar"):
    cls = "nx-price-card dass" if kind == "dass" else "nx-price-card solar"
    return f"""<div class="{cls}">
        <div class="nx-price-label">{label}</div>
        <div class="nx-price-value">{value}<span class="unit">{unit}</span></div>
        <div class="nx-price-foot">{foot}</div>
    </div>"""


def price_grid(cards: list[str]):
    st.markdown("<div class='nx-price-grid'>" + "".join(cards) + "</div>", unsafe_allow_html=True)


def stepbar(items: list[str]):
    html = "<div class='nx-stepbar'>"
    for i, item in enumerate(items, start=1):
        html += f"<div class='nx-step'><span>{i}</span>{item}</div>"
    html += "</div>"
    st.markdown(html, unsafe_allow_html=True)


def chart_heading(title: str, subtitle: str = "", kind: str = ""):
    cls = "nx-chart-title dass" if kind == "dass" else "nx-chart-title"
    st.markdown(f"<div class='{cls}'>{title}</div>", unsafe_allow_html=True)
    if subtitle:
        st.markdown(f"<div class='nx-chart-note'>{subtitle}</div>", unsafe_allow_html=True)


def style_chart(ch):
    return (ch.configure_axis(labelColor=MUTED, titleColor=MUTED, gridColor="#eef4f1",
                              domainColor="#dbe7e1", labelFontSize=11, titleFontSize=12)
              .configure_header(labelFontSize=13, labelFontWeight="bold", labelColor=INK,
                                titleColor=MUTED)
              .configure_view(strokeWidth=0))


# ---------------------------------------------------------------------------
# Fast data loading: xlsx -> parquet disk cache
# ---------------------------------------------------------------------------
@st.cache_data(show_spinner="Loading workbook summary…")
def load_summary_dataset(path_str: str, mtime: float) -> tuple[pd.DataFrame, pd.Series, pd.DataFrame]:
    """Fast path: read only workbook summary columns.

    This intentionally does NOT parse the hourly price/generation block. The normal
    Aurora workbook view only needs the already summarised monthly/annual values.
    Hourly data is loaded lazily only when a user uploads a custom forward curve.
    """
    path = Path(path_str)
    monthly_recs: list[dict] = []
    annual_recs: list[dict] = []
    share_vals: list[float] = []

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=2, max_col=26, values_only=True):
        # N:S = Year, Month, Uncurtailed capture, Curtailed capture, BESS revenue, TB4
        if row[13] is not None and row[14] is not None:
            monthly_recs.append({
                "year": int(row[13]),
                "month": int(row[14]),
                "capture_uncurtailed": pd.to_numeric(row[15], errors="coerce"),
                "capture_curtailed": pd.to_numeric(row[16], errors="coerce"),
                "rev_eur_mw": pd.to_numeric(row[17], errors="coerce"),
                "tb4": pd.to_numeric(row[18], errors="coerce"),
            })

        # U:Y = Year, Uncurtailed capture, Curtailed capture, BESS revenue, TB4
        if row[20] is not None:
            annual_recs.append({
                "year": int(row[20]),
                "capture_uncurtailed": pd.to_numeric(row[21], errors="coerce"),
                "capture_curtailed": pd.to_numeric(row[22], errors="coerce"),
                "rev_eur_mw": pd.to_numeric(row[23], errors="coerce"),
                "tb4": pd.to_numeric(row[24], errors="coerce"),
            })

        # Z = monthly solar shape / volume distribution. If absent, equal months.
        if row[25] is not None and len(share_vals) < 12:
            try:
                share_vals.append(float(row[25]))
            except Exception:
                pass
    wb.close()

    monthly = pd.DataFrame(monthly_recs).dropna(subset=["year", "month"])
    annual = pd.DataFrame(annual_recs).dropna(subset=["year"])
    if not monthly.empty:
        monthly[["year", "month"]] = monthly[["year", "month"]].astype(int)
    if not annual.empty:
        annual["year"] = annual["year"].astype(int)

    if share_vals and sum(share_vals[:12]) != 0:
        shares = pd.Series(np.array(share_vals[:12]) / np.sum(share_vals[:12]),
                           index=range(1, 13), name="share")
    else:
        shares = pd.Series(np.ones(12) / 12, index=range(1, 13), name="share")
    return monthly, annual, shares


@st.cache_data(show_spinner="Loading hourly data for custom curve…")
def load_hourly_dataset(path_str: str, mtime: float) -> pd.DataFrame:
    """Slow path, used only when a user uploads a custom curve."""
    path = Path(path_str)
    pq = path.with_suffix(".parquet")
    if pq.exists() and pq.stat().st_mtime >= mtime:
        return pd.read_parquet(pq)

    hourly_recs: list[dict] = []
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=2, max_col=6, values_only=True):
        if row[1] is not None and row[4] is not None:
            dt = pd.to_datetime(row[0], errors="coerce")
            if pd.notna(dt):
                hourly_recs.append({
                    "year": int(row[1]),
                    "month": int(row[2]),
                    "day": int(dt.day),
                    "hour": int(row[3]),
                    "price": float(row[4]),
                    "solar": 0.0 if row[5] is None else float(row[5]),
                })
    wb.close()
    hourly = pd.DataFrame(hourly_recs)
    if not hourly.empty:
        hourly[["year", "month", "day", "hour"]] = hourly[["year", "month", "day", "hour"]].astype(int)
        hourly[["price", "solar"]] = hourly[["price", "solar"]].astype(float)
        try:
            hourly.to_parquet(pq, index=False)
        except Exception:
            pass
    return hourly


def add_summary_helpers(df: pd.DataFrame, monthly_shares: pd.Series) -> pd.DataFrame:
    """Add columns previously rebuilt from hourly data, using summary-table inputs only."""
    out = df.copy()
    out["settled_share"] = 1.0
    out["neg_hours_gen_share"] = np.nan
    out["solar_mwh"] = out["month"].map(monthly_shares).fillna(1 / 12).astype(float)
    return out


def calendar_days_monthly(year_range: tuple[int, int]) -> pd.DataFrame:
    dates = pd.date_range(f"{year_range[0]}-01-01", f"{year_range[1]}-12-01", freq="MS")
    return pd.DataFrame({
        "year": dates.year.astype(int),
        "month": dates.month.astype(int),
        "ndays": dates.days_in_month.astype(int),
    })


def calendar_days_annual(year_range: tuple[int, int]) -> pd.DataFrame:
    years = np.arange(year_range[0], year_range[1] + 1)
    return pd.DataFrame({"year": years.astype(int),
                         "ndays": [366 if pd.Timestamp(int(y), 12, 31).is_leap_year else 365 for y in years]})


@st.cache_data(show_spinner=False)
def typical_solar_shape(hourly: pd.DataFrame) -> pd.DataFrame:
    hist = hourly[hourly["year"] <= 2025]
    return (hist.groupby(["month", "hour"])["solar"].mean()
                .reset_index().rename(columns={"solar": "solar_shape"}))


# ---------------------------------------------------------------------------
# BESS dispatch — vectorised (complete 24h days) + loop fallback for odd days
# ---------------------------------------------------------------------------
CH_W = np.array([1.0, 1.0, 1.0, 1.0, E_CHARGE_GRID - 4.0])   # 5 cheapest, partial last
DIS_W = np.array([E_DISCHARGE - 3.0, 1.0, 1.0, 1.0])          # 4 priciest, partial first


def _dispatch_day_loop(p: np.ndarray) -> tuple[float, float]:
    n = len(p)
    if n < 10:
        return np.nan, np.nan
    order = np.argsort(p)
    ch = np.zeros(n)
    rem = E_CHARGE_GRID
    for i in order:
        t = min(BESS_POWER_MW, rem)
        ch[i] = t
        rem -= t
        if rem <= 1e-9:
            break
    di = np.zeros(n)
    rem = E_DISCHARGE
    for i in order[::-1]:
        t = min(BESS_POWER_MW, rem)
        di[i] = t
        rem -= t
        if rem <= 1e-9:
            break
    srt = np.sort(p)
    return float((di * p).sum() - (ch * p).sum()), float(srt[-4:].mean() - srt[:4].mean())


@st.cache_data(show_spinner="Optimising BESS dispatch…")
def bess_daily_results(hourly: pd.DataFrame) -> pd.DataFrame:
    df = hourly.sort_values(["year", "month", "day", "hour"])
    sizes = df.groupby(["year", "month", "day"], sort=True).size()
    full_keys = sizes[sizes == 24].index
    odd_keys = sizes[sizes != 24].index

    df_idx = df.set_index(["year", "month", "day"])
    recs = []

    if len(full_keys):
        block = df_idx.loc[full_keys, "price"].to_numpy().reshape(-1, 24)
        s = np.sort(block, axis=1)
        rev = s[:, -4:] @ DIS_W - s[:, :5] @ CH_W
        tb4 = s[:, -4:].mean(axis=1) - s[:, :4].mean(axis=1)
        keys = np.array(list(full_keys))
        recs.append(pd.DataFrame({"year": keys[:, 0], "month": keys[:, 1],
                                  "day": keys[:, 2], "rev_eur_mw": rev, "tb4": tb4}))
    for key in odd_keys:
        rev, tb4 = _dispatch_day_loop(df_idx.loc[[key], "price"].to_numpy(float))
        recs.append(pd.DataFrame({"year": [key[0]], "month": [key[1]], "day": [key[2]],
                                  "rev_eur_mw": [rev], "tb4": [tb4]}))
    return pd.concat(recs, ignore_index=True).dropna()


@st.cache_data(show_spinner=False)
def monthly_capture(hourly: pd.DataFrame, settle_at_nonpositive: bool) -> pd.DataFrame:
    df = hourly.copy()
    df["w"] = df["solar"]
    df["w_settle"] = df["solar"] if settle_at_nonpositive \
        else np.where(df["price"] > 0, df["solar"], 0.0)
    df["pw"] = df["price"] * df["w_settle"]
    g = df.groupby(["year", "month"]).agg(pw=("pw", "sum"), ws=("w_settle", "sum"),
                                          w=("w", "sum")).reset_index()
    g["capture"] = np.where(g["ws"] > 0, g["pw"] / g["ws"], np.nan)
    g["settled_share"] = np.where(g["w"] > 0, g["ws"] / g["w"], np.nan)
    g["neg_hours_gen_share"] = 1 - g["settled_share"]
    g["solar_mwh"] = g["w"]  # relative units, used only for shaping volumes
    return g[["year", "month", "capture", "settled_share",
              "neg_hours_gen_share", "solar_mwh"]]


# ---------------------------------------------------------------------------
# Uploaded forward curve
# ---------------------------------------------------------------------------
def parse_uploaded_curve(uploaded) -> pd.DataFrame | None:
    try:
        raw = (pd.read_excel(uploaded) if uploaded.name.lower().endswith((".xlsx", ".xls"))
               else pd.read_csv(uploaded))
    except Exception as exc:  # noqa: BLE001
        st.error(f"Could not read the file: {exc}")
        return None
    dt_col, price_col = None, None
    for c in raw.columns:
        parsed = pd.to_datetime(raw[c], errors="coerce")
        if parsed.notna().mean() > 0.9:
            dt_col, raw[c] = c, parsed
            break
    for c in raw.columns:
        if c == dt_col:
            continue
        vals = pd.to_numeric(raw[c], errors="coerce")
        if vals.notna().mean() > 0.9:
            price_col, raw[c] = c, vals
            break
    if dt_col is None or price_col is None:
        st.error("The file needs one datetime column and one numeric hourly price column.")
        return None
    out = pd.DataFrame({
        "year": raw[dt_col].dt.year, "month": raw[dt_col].dt.month,
        "day": raw[dt_col].dt.day, "hour": raw[dt_col].dt.hour + 1,
        "price": raw[price_col].astype(float)}).dropna()
    out[["year", "month", "day"]] = out[["year", "month", "day"]].astype(int)
    return out


def curve_template_bytes() -> bytes:
    idx = pd.date_range("2027-01-01 00:00", "2027-12-31 23:00", freq="h")
    tmpl = pd.DataFrame({"Datetime (hourly)": idx, "Price (€/MWh)": np.nan})
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xw:
        tmpl.to_excel(xw, index=False, sheet_name="forward_curve")
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Chart builders — responsive, no faceting, labels centred inside bars
# ---------------------------------------------------------------------------
def _x_encoding(granularity: str):
    if granularity == "Annual":
        return alt.X("period:N", title=None, sort=None, axis=alt.Axis(labelAngle=0))
    return alt.X("date:T", title=None,
                 axis=alt.Axis(format="%b", labelAngle=0, tickCount=24,
                               labelOverlap=True, labelFontSize=10, labelPadding=8))


def _monthly_year_guides(df: pd.DataFrame):
    years = (df[["year"]].drop_duplicates().sort_values("year").reset_index(drop=True))
    if years.empty:
        return None, None, None
    years["x1"] = pd.to_datetime(dict(year=years["year"], month=1, day=1))
    years["x2"] = pd.to_datetime(dict(year=years["year"] + 1, month=1, day=1))
    years["mid"] = years["x1"] + (years["x2"] - years["x1"]) / 2
    years["shade"] = np.where(years.index % 2 == 0, "A", "B")

    bands = alt.Chart(years).mark_rect(opacity=0.06).encode(
        x="x1:T", x2="x2:T",
        color=alt.Color("shade:N", scale=alt.Scale(domain=["A", "B"], range=["#f4f7f6", "#ffffff"]), legend=None)
    )
    rules = alt.Chart(years).mark_rule(color="#d5ddd9", strokeDash=[3, 3]).encode(x="x1:T")
    labels = alt.Chart(years).mark_text(
        baseline="top", align="center", dy=0, fontSize=10, color=MUTED, fontWeight="bold"
    ).encode(
        x="mid:T", y=alt.value(6), text="year:O"
    )
    return bands, rules, labels


def _bar_label(v: float, suffix: str = "") -> str:
    if pd.isna(v):
        return ""
    av = abs(v)
    if av >= 1_000_000:
        return f"{v/1_000_000:+.1f}M{suffix}"
    if av >= 1_000:
        return f"{v/1_000:+.0f}k{suffix}"
    return f"{v:+.1f}{suffix}"


def settlement_chart(df: pd.DataFrame, ycol: str, ytitle: str,
                     pos_lbl: str, neg_lbl: str, pos_c: str, neg_c: str,
                     granularity: str, title: str, tooltips: list | None = None,
                     height: int = 430, label_suffix: str = ""):
    df = df.copy()
    df["sign"] = np.where(df[ycol] >= 0, pos_lbl, neg_lbl)
    df["fill_color"] = np.where(df[ycol] >= 0, pos_c, neg_c)
    df["bar_lbl"] = df[ycol].map(lambda v: _bar_label(v, label_suffix))
    tt = tooltips or []

    # Back to bars for Monthly, but with year banding and black labels outside bars.
    # Use explicit per-row hex colours so positive bars are clearly green and negative bars clearly orange.
    color = alt.Color(
        "fill_color:N",
        scale=None,
        legend=None,
    )
    xenc = _x_encoding(granularity)
    base = alt.Chart(df)

    zero = base.mark_rule(color="#adc5bc", strokeWidth=1.2).encode(y=alt.datum(0))
    bars = base.mark_bar(
        cornerRadiusTopLeft=4,
        cornerRadiusTopRight=4,
        cornerRadiusBottomLeft=4,
        cornerRadiusBottomRight=4,
        opacity=0.95,
        size=8 if granularity == "Monthly" else None,
        stroke="#ffffff",
        strokeWidth=0.6,
    ).encode(
        x=xenc,
        y=alt.Y(f"{ycol}:Q", title=ytitle),
        color=color,
        tooltip=tt,
    )

    # Black labels, outside the bar, with a small white outline so they remain readable
    # even when they sit close to the coloured bar or gridline.
    label_size = 9 if granularity == "Monthly" else 11
    pos_labels = (base.transform_filter(f"datum.{ycol} >= 0")
        .mark_text(
            dy=-7,
            fontSize=label_size,
            fontWeight="bold",
            color=INK,
            stroke="white",
            strokeWidth=2,
            strokeOpacity=0.85,
        )
        .encode(x=xenc, y=alt.Y(f"{ycol}:Q"), text="bar_lbl:N"))
    neg_labels = (base.transform_filter(f"datum.{ycol} < 0")
        .mark_text(
            dy=12,
            fontSize=label_size,
            fontWeight="bold",
            color=INK,
            stroke="white",
            strokeWidth=2,
            strokeOpacity=0.85,
        )
        .encode(x=xenc, y=alt.Y(f"{ycol}:Q"), text="bar_lbl:N"))

    layers = []
    if granularity == "Monthly":
        bands, rules, year_labels = _monthly_year_guides(df)
        layers.extend([bands, rules, year_labels])
    layers.extend([bars, zero, pos_labels, neg_labels])

    chart_height = max(height, 450) if granularity == "Monthly" else height
    return alt.layer(*layers).properties(
        height=chart_height,
        title=alt.TitleParams(title, anchor="start", fontSize=15, color=INK),
    )

def market_vs_contract_chart(df: pd.DataFrame, market_col: str, contract_col: str,
                             ytitle: str, title: str, market_name: str,
                             contract_name: str, granularity: str,
                             tooltips: list | None = None, height: int = 390):
    df = df.copy()
    df["market_lbl"] = df[market_col].map(lambda v: "" if pd.isna(v) else f"{v:.0f}")
    show_labels = (granularity == "Annual") or (len(df) <= 96)
    xenc = _x_encoding(granularity)
    tt = tooltips or []
    base = alt.Chart(df)
    bars = base.mark_bar(cornerRadiusTopLeft=5, cornerRadiusTopRight=5, opacity=0.88).encode(
        x=xenc,
        y=alt.Y(f"{market_col}:Q", title=ytitle),
        color=alt.value(GREEN),
        tooltip=tt,
    )
    line = base.mark_line(point={"filled": True, "size": 70}, strokeWidth=3, color=ORANGE).encode(
        x=xenc,
        y=alt.Y(f"{contract_col}:Q", title=ytitle),
        tooltip=tt,
    )
    layers = []
    if granularity == "Monthly":
        bands, rules, year_labels = _monthly_year_guides(df)
        layers.extend([bands, rules, year_labels])
    layers.extend([bars, line])
    if show_labels:
        bar_labels = base.mark_text(dy=-8, fontSize=10, fontWeight="bold", color=INK).encode(
            x=xenc, y=alt.Y(f"{market_col}:Q"), text="market_lbl:N")
        layers.append(bar_labels)
    return alt.layer(*layers).properties(
        height=height,
        title=alt.TitleParams(f"{title} · bars = {market_name}, line = {contract_name}",
                              anchor="start", fontSize=15, color=INK))


# ---------------------------------------------------------------------------
# Load data
# ---------------------------------------------------------------------------
data_path = resolve_data_file()
if data_path is None:
    st.error("Data file `PV___BESS_settlement_app.xlsx` not found in `data/`.")
    st.stop()

monthly_summary_base, annual_summary_base, monthly_shares = load_summary_dataset(
    str(data_path), data_path.stat().st_mtime)
HIST_MAX_YEAR = 2026

# ---------------------------------------------------------------------------
# Global controls
# ---------------------------------------------------------------------------
with st.container(border=True):
    top1, top2, top4 = st.columns([1.0, 1.4, 1.6])
    with top1:
        granularity = st.radio("View", ["Annual", "Monthly"], horizontal=True)
    with top2:
        year_range = st.slider("Period", 2021, 2040, (2024, 2030))
    with top4:
        curve_source = st.radio("Forward curve (≥ 2027)",
                                ["Aurora (workbook)", "Upload my own curve"],
                                horizontal=True)

uploaded_curve = None
if curve_source.startswith("Upload"):
    up_col1, up_col2 = st.columns([2, 1])
    with up_col1:
        up = st.file_uploader("Hourly forward price curve (CSV/XLSX: datetime + €/MWh)",
                              type=["csv", "xlsx", "xls"])
        if up is not None:
            uploaded_curve = parse_uploaded_curve(up)
    with up_col2:
        st.download_button(
            "⬇️ Template (XLSX)", data=curve_template_bytes(),
            file_name="forward_curve_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

use_uploaded_curve = uploaded_curve is not None and not uploaded_curve.empty

if use_uploaded_curve:
    hourly_base = load_hourly_dataset(str(data_path), data_path.stat().st_mtime)
    solar_shape = typical_solar_shape(hourly_base)
    fwd_years = sorted(uploaded_curve["year"].unique())
    hist_part = hourly_base[hourly_base["year"] <= HIST_MAX_YEAR].copy()
    fwd = uploaded_curve.merge(solar_shape, on=["month", "hour"], how="left")
    fwd["solar"] = fwd["solar_shape"].fillna(0.0)
    fwd = fwd[["year", "month", "day", "hour", "price", "solar"]]
    hourly = pd.concat([hist_part, fwd], ignore_index=True)
    hourly = hourly[(hourly.year >= year_range[0]) & (hourly.year <= year_range[1])]
    st.info(f"Using **your uploaded curve** for {fwd_years[0]}–{fwd_years[-1]} "
            f"({len(uploaded_curve):,} hours). Solar capture and BESS dispatch are "
            "recomputed from the uploaded prices.")
else:
    hourly = None
    st.caption("Fast mode: using only workbook summary tables for captured prices, BESS revenues and TB4. The hourly sheet is not loaded unless a user uploads a curve.")

stepbar(["Choose annual or monthly view", "Set tenor and forward curve", "Review Solar / DASS charts and export to PDF"])

ms = None
if use_uploaded_curve:
    if hourly is None or hourly.empty:
        st.warning("No data in the selected period.")
        st.stop()
    bess_day = bess_daily_results(hourly)
    bess_m = (bess_day.groupby(["year", "month"])
              .agg(rev_eur_mw=("rev_eur_mw", "sum"), tb4=("tb4", "mean")).reset_index())
    annual_summary_bess = None
    ndays_m = hourly.groupby(["year", "month"])["day"].nunique().reset_index(name="ndays")
    ndays_y = hourly.groupby("year")[["month", "day"]].apply(
        lambda x: x.drop_duplicates().shape[0]).reset_index(name="ndays")
else:
    ms = monthly_summary_base[(monthly_summary_base.year >= year_range[0]) &
                              (monthly_summary_base.year <= year_range[1])].copy()
    if ms.empty:
        st.warning("No summary data in the selected period.")
        st.stop()
    bess_m = ms[["year", "month", "rev_eur_mw", "tb4"]].copy()
    annual_summary_bess = annual_summary_base[(annual_summary_base.year >= year_range[0]) &
                                              (annual_summary_base.year <= year_range[1])].copy()
    ndays_m = calendar_days_monthly(year_range)
    ndays_y = calendar_days_annual(year_range)

# A real date axis makes Monthly charts responsive and avoids squeezing mini-facets.
for _df in [bess_m]:
    if _df is not None and not _df.empty:
        _df["date"] = pd.to_datetime(dict(year=_df["year"], month=_df["month"], day=1))
        _df["period"] = _df["year"].astype(str) + "-" + _df["month"].map(lambda m: f"{m:02d}")

# ======================================================================
# 1) SOLAR PV PPA
# ======================================================================
start_module_wrap("solar")
module_banner("Solar PPA simulation", "Solar PPA settlement view — captured price, strike / floor and cash settlement.", "Solar PV PPA")
st.markdown("<div class='nx-section-card-note'>Solar section · all charts and KPIs below refer to the selected Solar PPA configuration.</div>", unsafe_allow_html=True)
section("Solar PV PPA", GREEN,
        "First chart: captured solar price in the market vs the contract price. "
        "Second chart: settlement to the offtaker under the selected convention.",
        "☀️ Offtaker receives when green")

with st.container(border=True):
    c1, c2 = st.columns([1.1, 1.0])
    with c1:
        ppa_type = st.radio("PPA structure", ["Fixed for floating", "Floor + discount"],
                            horizontal=True)
        ppa_volume_gwh = st.number_input("PPA volume (GWh/yr)", 1.0, 5000.0, 100.0, 10.0)
    with c2:
        if ppa_type == "Fixed for floating":
            strike = st.slider("PPA strike price (€/MWh)", 0.0, 150.0, 30.0, 0.5)
            floor, discount = 0.0, 0.0
        else:
            strike = 0.0
            floor = st.slider("Floor (€/MWh)", 0.0, 150.0, 30.0, 0.5)
            discount = st.slider("Discount to market (€/MWh)", 0.0, 80.0, 5.0, 0.5,
                                 help="Contract price = max(floor, captured price − discount).")
        settle_nonpos = st.toggle(
            "Settle at 0 / negative prices", value=True,
            help="ON → use uncurtailed workbook capture. OFF → use curtailed workbook capture and exclude ≤0 €/MWh hours from settlement when hourly data is available.")

if ppa_type == "Fixed for floating":
    price_grid([
        price_card("PPA strike price", f"{strike:.1f}", "€/MWh", "Fixed strike used as the orange line in the chart."),
        price_card("Settlement logic", "Capture − Strike", "", "Positive means payment to the offtaker."),
        price_card("Contracted volume", f"{ppa_volume_gwh:,.0f}", "GWh/yr", "Annual PPA volume shaped by the solar profile."),
    ])
else:
    price_grid([
        price_card("PPA floor / strike", f"{floor:.1f}", "€/MWh", "This is the fixed orange line. It does not move with captured price."),
        price_card("Discount to market", f"{discount:.1f}", "€/MWh", "Effective price = max(floor, captured price − discount)."),
        price_card("Contracted volume", f"{ppa_volume_gwh:,.0f}", "GWh/yr", "Annual PPA volume shaped by the solar profile."),
    ])

capture_col = "capture_uncurtailed" if settle_nonpos else "capture_curtailed"
if use_uploaded_curve:
    cap_m = monthly_capture(hourly, settle_nonpos)
    annual_summary_solar = None
else:
    cap_m = ms[["year", "month", capture_col]].rename(columns={capture_col: "capture"}).copy()
    cap_m = add_summary_helpers(cap_m, monthly_shares)
    annual_summary_solar = annual_summary_base[(annual_summary_base.year >= year_range[0]) &
                                               (annual_summary_base.year <= year_range[1])].copy()

if cap_m is not None and not cap_m.empty:
    cap_m["date"] = pd.to_datetime(dict(year=cap_m["year"], month=cap_m["month"], day=1))
    cap_m["period"] = cap_m["year"].astype(str) + "-" + cap_m["month"].map(lambda m: f"{m:02d}")

ppa = cap_m.copy()
ppa["yr_solar"] = ppa.groupby("year")["solar_mwh"].transform("sum")
ppa["volume_mwh"] = ppa_volume_gwh * 1000.0 * ppa["solar_mwh"] / ppa["yr_solar"].clip(lower=1e-9)
ppa["settled_mwh"] = ppa["volume_mwh"] * ppa["settled_share"].fillna(1.0)
if ppa_type == "Fixed for floating":
    ppa["contract_price"] = strike
    ppa["strike_floor_price"] = strike
else:
    ppa["contract_price"] = np.maximum(floor, ppa["capture"] - discount)
    ppa["strike_floor_price"] = floor
ppa["settle_eur_mwh"] = ppa["capture"] - ppa["contract_price"]
ppa["settle_eur"] = ppa["settle_eur_mwh"] * ppa["settled_mwh"]

if granularity == "Annual":
    if (not use_uploaded_curve) and annual_summary_solar is not None:
        ppa_view = annual_summary_solar[["year", capture_col]].rename(columns={capture_col: "capture"}).copy()
        vol_y = ppa.groupby("year").agg(settled_mwh=("settled_mwh", "sum"),
                                        volume_mwh=("volume_mwh", "sum"),
                                        neg_hours_gen_share=("neg_hours_gen_share", "mean")).reset_index()
        ppa_view = ppa_view.merge(vol_y, on="year", how="left")
        if ppa_type == "Fixed for floating":
            ppa_view["contract_price"] = strike
            ppa_view["strike_floor_price"] = strike
        else:
            ppa_view["contract_price"] = np.maximum(floor, ppa_view["capture"] - discount)
            ppa_view["strike_floor_price"] = floor
        ppa_view["settle_eur_mwh"] = ppa_view["capture"] - ppa_view["contract_price"]
        ppa_view["settle_eur"] = ppa_view["settle_eur_mwh"] * ppa_view["settled_mwh"].fillna(0)
    else:
        g = ppa.groupby("year")
        ppa_view = pd.DataFrame({
            "settle_eur": g["settle_eur"].sum(),
            "settled_mwh": g["settled_mwh"].sum(),
            "volume_mwh": g["volume_mwh"].sum(),
            "capture": g.apply(lambda x: (x["capture"] * x["settled_mwh"]).sum()
                               / max(x["settled_mwh"].sum(), 1e-9)),
            "contract_price": g.apply(lambda x: (x["contract_price"] * x["settled_mwh"]).sum()
                                      / max(x["settled_mwh"].sum(), 1e-9)),
            "neg_hours_gen_share": g["neg_hours_gen_share"].mean(),
        }).reset_index()
        ppa_view["settle_eur_mwh"] = ppa_view["settle_eur"] / ppa_view["settled_mwh"].clip(lower=1e-9)
        ppa_view["strike_floor_price"] = strike if ppa_type == "Fixed for floating" else floor
    ppa_view["period"] = ppa_view["year"].astype(str)
    ppa_view["month"] = 1
else:
    ppa_view = ppa.copy()

if "strike_floor_price" not in ppa_view.columns:
    ppa_view["strike_floor_price"] = strike if ppa_type == "Fixed for floating" else floor
if "date" not in ppa_view.columns:
    ppa_view["date"] = pd.to_datetime(dict(year=ppa_view["year"], month=ppa_view["month"], day=1))

avg_settle_mwh = ppa_view["settle_eur"].sum() / max(ppa_view["settled_mwh"].sum(), 1e-9)
avg_capture = (ppa_view["capture"] * ppa_view["settled_mwh"].fillna(0)).sum() / max(ppa_view["settled_mwh"].sum(), 1e-9)
tone_ppa = "pos" if avg_settle_mwh >= 0 else "neg"

k1, k2, k3, k4 = st.columns(4)
kpi(k1, "Avg settlement to offtaker", f"{avg_settle_mwh:+.1f}", "€/MWh",
    "Per settled MWh, selected period", tone_ppa)
kpi(k2, "Cumulative settlement", f"{ppa_view['settle_eur'].sum()/1e6:+.2f}", "M€",
    f"{ppa_volume_gwh:,.0f} GWh/yr · {'settles ≤0€' if settle_nonpos else 'no settle ≤0€'}",
    tone_ppa)
kpi(k3, "Avg captured price", f"{avg_capture:.1f}", "€/MWh",
    "Solar-weighted market price", "neu")
neg_share = ppa["neg_hours_gen_share"].dropna()
neg_share_txt = "n/a" if neg_share.empty else f"{100*neg_share.mean():.1f}"
kpi(k4, "Generation in ≤0 € hours", neg_share_txt, "%",
    "Share of solar volume at non-positive prices", "neu")

st.markdown("")
chart_heading("Strike / floor price vs solar captured price", "Orange line = fixed strike / floor. Green bars = solar captured market price. Monthly charts include year banding so each month is easier to place within its year.")
tt_ppa = [alt.Tooltip("period:N", title="Period"),
          alt.Tooltip("capture:Q", title="Captured €/MWh", format=".1f"),
          alt.Tooltip("strike_floor_price:Q", title="Strike / floor €/MWh", format=".1f"),
          alt.Tooltip("contract_price:Q", title="Effective contract €/MWh", format=".1f"),
          alt.Tooltip("settle_eur_mwh:Q", title="Settlement €/MWh", format="+.1f"),
          alt.Tooltip("settle_eur:Q", title="Settlement €", format=",.0f")]
st.altair_chart(style_chart(market_vs_contract_chart(
    ppa_view, "capture", "strike_floor_price", "€/MWh",
    "Strike / floor price vs solar captured price",
    "solar captured price", "strike / floor", granularity, tooltips=tt_ppa)),
    use_container_width=True)

chart_heading("Settlement in €/MWh", "Positive values = payment to the offtaker; negative values = payment by the offtaker. Bars are green when positive and orange when negative.")
st.altair_chart(style_chart(settlement_chart(
    ppa_view, "settle_eur_mwh", "Settlement to offtaker (€/MWh)",
    "Offtaker receives", "Offtaker pays", GREEN, ORANGE, granularity,
    "PPA settlement per MWh — positive means payment to offtaker",
    tooltips=tt_ppa)), use_container_width=True)

chart_heading("Settlement in €", "Total cash settlement using the selected PPA volume. Bars are green when positive and orange when negative.")
st.altair_chart(style_chart(settlement_chart(
    ppa_view, "settle_eur", "Settlement to offtaker (€)",
    "Offtaker receives", "Offtaker pays", GREEN, ORANGE, granularity,
    f"PPA settlement in € — {ppa_volume_gwh:,.0f} GWh/yr shaped on solar profile",
    tooltips=[alt.Tooltip("period:N"),
              alt.Tooltip("settle_eur:Q", title="Settlement €", format=",.0f")],
    height=320)), use_container_width=True)

with st.expander("PPA settlement table"):
    tbl = (ppa if granularity == "Monthly" else ppa_view)
    cols = ["period", "capture", "strike_floor_price", "contract_price", "settle_eur_mwh",
            "volume_mwh", "settled_mwh", "settle_eur", "neg_hours_gen_share"]
    st.dataframe(tbl[cols].rename(columns={
        "capture": "Captured €/MWh", "strike_floor_price": "Strike / Floor €/MWh", "contract_price": "Effective Contract €/MWh",
        "settle_eur_mwh": "Settlement €/MWh", "volume_mwh": "Volume MWh",
        "settled_mwh": "Settled MWh", "settle_eur": "Settlement €",
        "neg_hours_gen_share": "% gen in ≤0€ h"}).round(2),
        use_container_width=True, hide_index=True)

end_module_wrap()

# ======================================================================
# 2) BESS DASS
# ======================================================================
start_module_wrap("dass")
module_banner("Battery spread swap module", "Day-ahead BESS revenue versus DASS strike and resulting settlement.", "BESS DASS", "dass")
st.markdown("<div class='nx-section-card-note dass'>BESS section · all charts and KPIs below refer to the selected DASS configuration.</div>", unsafe_allow_html=True)
section("BESS Day-Ahead Spread Swap (DASS)", DASS_GREEN,
        "First chart: realised / forecast BESS market revenue vs the DASS strike. "
        "Second chart: the settlement after netting market revenue against the strike.",
        "🔋 Buyer receives when green", "dass")

with st.container(border=True):
    d1, d2, d3 = st.columns([1.4, 1.0, 1.4])
    with d1:
        dass_strike = st.slider("DASS strike (k€/MW·yr)", 0.0, 200.0, 70.0, 0.5)
    with d2:
        dass_mw = st.number_input("Contracted BESS capacity (MW)", 0.5, 1000.0, 1.0, 0.5)
    with d3:
        eq_spread = dass_strike * 1000 / (365 * BESS_CAPACITY_MWH * RTE)
        st.markdown(
            f"<div style='color:{MUTED};font-size:.87rem;padding-top:26px;'>"
            f"Strike ≈ <b>{eq_spread:.1f} €/MWh</b> TB4 spread equivalent "
            f"(strike ÷ 365 d × 4 h × 1 c/d × 0.856 RTE).</div>",
            unsafe_allow_html=True)

st.markdown(
    f"""<div style="background:{DASS_GREEN_SOFT};border:1px solid #c8ecd7;border-radius:12px;
    padding:10px 16px;margin:6px 0 4px 0;color:{DASS_GREEN};font-size:.85rem;font-weight:600;">
    ⚙️ BESS assumptions &nbsp;·&nbsp; 1 MW / 4 MWh &nbsp;·&nbsp; max 1 cycle/day
    &nbsp;·&nbsp; η<sub>ch</sub> = η<sub>dis</sub> = 0.925 → RTE ≈ 0.856 &nbsp;·&nbsp;
    charges {E_CHARGE_GRID:.2f} MWh / discharges {E_DISCHARGE:.2f} MWh per cycle
    &nbsp;·&nbsp; undegraded &nbsp;·&nbsp; forward prices nominal</div>""",
    unsafe_allow_html=True)

price_grid([
    price_card("DASS strike", f"{dass_strike:.1f}", "k€/MW·yr", "Annual fixed strike. Orange line in the BESS chart.", "dass"),
    price_card("TB4 equivalent", f"{eq_spread:.1f}", "€/MWh", "Approximate spread equivalent using 4h, 1 cycle/day and 0.856 RTE.", "dass"),
    price_card("Contracted capacity", f"{dass_mw:,.1f}", "MW", "Used to convert €/MW settlement into total cash settlement.", "dass"),
])

bess = bess_m.copy().merge(ndays_m, on=["year", "month"], how="left")
bess["strike_eur_mw"] = dass_strike * 1000.0 * bess["ndays"] / 365.0
bess["settle_eur_mw"] = bess["rev_eur_mw"] - bess["strike_eur_mw"]
bess["settle_eur"] = bess["settle_eur_mw"] * dass_mw
bess["rev_keur_mw"] = bess["rev_eur_mw"] / 1000.0
bess["strike_keur_mw"] = bess["strike_eur_mw"] / 1000.0

if granularity == "Annual":
    if (not use_uploaded_curve) and annual_summary_bess is not None:
        bess_view = annual_summary_bess[["year", "rev_eur_mw", "tb4"]].copy()
        bess_view = bess_view.merge(ndays_y, on="year", how="left")
        bess_view["strike_eur_mw"] = dass_strike * 1000.0 * bess_view["ndays"] / 365.0
        bess_view["settle_eur_mw"] = bess_view["rev_eur_mw"] - bess_view["strike_eur_mw"]
        bess_view["settle_eur"] = bess_view["settle_eur_mw"] * dass_mw
    else:
        bess_view = (bess.groupby("year")
                     .agg(rev_eur_mw=("rev_eur_mw", "sum"),
                          strike_eur_mw=("strike_eur_mw", "sum"),
                          settle_eur_mw=("settle_eur_mw", "sum"),
                          settle_eur=("settle_eur", "sum"),
                          tb4=("tb4", "mean"), ndays=("ndays", "sum")).reset_index())
    bess_view["period"] = bess_view["year"].astype(str)
    bess_view["month"] = 1
else:
    bess_view = bess.copy()

if "date" not in bess_view.columns:
    bess_view["date"] = pd.to_datetime(dict(year=bess_view["year"], month=bess_view["month"], day=1))
bess_view["settle_keur_mw"] = bess_view["settle_eur_mw"] / 1000.0
bess_view["rev_keur_mw"] = bess_view["rev_eur_mw"] / 1000.0
bess_view["strike_keur_mw"] = bess_view["strike_eur_mw"] / 1000.0

avg_settle = bess_view["settle_eur_mw"].mean() / 1000
tone_dass = "pos" if avg_settle >= 0 else "neg"
k1, k2, k3, k4 = st.columns(4)
kpi(k1, "Avg settlement to buyer", f"{avg_settle:+.1f}",
    f"k€/MW·{'yr' if granularity == 'Annual' else 'mo'}",
    f"Strike {dass_strike:.0f} k€/MW·yr ≈ {eq_spread:.0f} €/MWh TB4", tone_dass)
kpi(k2, "Cumulative settlement", f"{bess_view['settle_eur'].sum()/1e6:+.2f}", "M€",
    f"{dass_mw:.0f} MW contracted", tone_dass)
kpi(k3, "Avg BESS revenue", f"{bess_view['rev_eur_mw'].mean()/1000:.1f}",
    f"k€/MW·{'yr' if granularity == 'Annual' else 'mo'}",
    "1 MW/4 MWh · 1 cycle/day · RTE 0.856", "pur")
kpi(k4, "Avg market TB4 spread", f"{bess_view['tb4'].mean():.1f}", "€/MWh",
    "Mean of daily top-4 − bottom-4 prices", "neu")

st.markdown("")
chart_heading("DASS strike vs BESS market revenue", "Green bars = BESS market revenue; orange line = DASS strike.", "dass")
tt_dass = [alt.Tooltip("period:N", title="Period"),
           alt.Tooltip("rev_eur_mw:Q", title="BESS revenue €/MW", format=",.0f"),
           alt.Tooltip("strike_eur_mw:Q", title="DASS strike €/MW", format=",.0f"),
           alt.Tooltip("settle_keur_mw:Q", title="Settlement k€/MW", format="+.1f"),
           alt.Tooltip("settle_eur:Q", title="Settlement €", format=",.0f"),
           alt.Tooltip("tb4:Q", title="TB4 spread €/MWh", format=".1f")]

st.altair_chart(style_chart(market_vs_contract_chart(
    bess_view, "rev_keur_mw", "strike_keur_mw", "k€/MW",
    "BESS market revenue vs DASS strike",
    "BESS revenue", "DASS strike", granularity, tooltips=tt_dass)),
    use_container_width=True)

chart_heading("Settlement in k€/MW", "Positive values = payment to the swap buyer; negative values = payment by the swap buyer. Bars are green when positive and orange when negative.", "dass")
st.altair_chart(style_chart(settlement_chart(
    bess_view, "settle_keur_mw", "Settlement to buyer (k€/MW)",
    "Buyer receives", "Buyer pays", GREEN, ORANGE, granularity,
    "DASS settlement — positive means payment to swap buyer",
    tooltips=tt_dass)), use_container_width=True)

chart_heading("Settlement in €", "Total cash settlement using the selected contracted BESS MW. Bars are green when positive and orange when negative.", "dass")
st.altair_chart(style_chart(settlement_chart(
    bess_view, "settle_eur", "Settlement to buyer (€)",
    "Buyer receives", "Buyer pays", GREEN, ORANGE, granularity,
    f"DASS settlement in € — {dass_mw:.0f} MW contracted",
    tooltips=[alt.Tooltip("period:N"),
              alt.Tooltip("settle_eur:Q", title="Settlement €", format=",.0f")],
    height=320)), use_container_width=True)

with st.expander("DASS settlement table"):
    st.dataframe(
        bess_view[["period", "rev_eur_mw", "strike_eur_mw", "settle_eur_mw",
                   "settle_eur", "tb4"]]
        .rename(columns={"rev_eur_mw": "BESS revenue €/MW",
                         "strike_eur_mw": "Strike €/MW",
                         "settle_eur_mw": "Settlement €/MW",
                         "settle_eur": "Settlement €",
                         "tb4": "TB4 spread €/MWh"}).round(1),
        use_container_width=True, hide_index=True)

end_module_wrap()

# ---------------------------------------------------------------------------
# Methodology / data lineage
# ---------------------------------------------------------------------------
with st.expander("ℹ️ Where every number comes from (data lineage & reconciliation)"):
    st.markdown(
        f"""
- **Single data source**: `data/PV___BESS_settlement_app.xlsx` — hourly day-ahead price
  (2021 – Jun 2026 OMIE outturn; 2027 – 2040 Aurora **nominal**), hourly solar profile,
  and workbook summary tables.
- **Aurora/workbook curve selected**: the app reads the workbook's monthly / annual
  summary figures directly: uncurtailed or curtailed solar captured price, BESS revenue
  and TB4. The settlement is then calculated from those figures and from the contract
  inputs selected in the app.
- **User-uploaded curve selected**: the app recomputes captured price and BESS revenue
  from the uploaded hourly prices. BESS revenue = one optimal cycle/day buying
  {E_CHARGE_GRID:.2f} MWh in the cheapest hours and selling {E_DISCHARGE:.2f} MWh in the
  most expensive (η = 0.925/0.925 → RTE ≈ 0.856, undegraded). **TB4** = daily mean of
  4 highest prices − mean of 4 lowest prices, averaged per period.
- **PPA Floor + Discount**: discount is in **€/MWh**. Orange line = floor. Effective settlement price = max(floor,
  captured price − discount).
- **Sign convention**: positive = payment **to the offtaker / swap buyer**.
  2026 figures are year-to-date where the workbook only contains data through June.
  Values indicative, pre-fees.
"""
    )

# ---------------------------------------------------------------------------
# Contact — questions about the settlements go to Nexwell Power
# ---------------------------------------------------------------------------
CONTACT_TO = os.getenv("CONTACT_TO", "mmoreno@nexwellpower.com")


def _smtp_config() -> dict | None:
    """SMTP credentials from st.secrets['smtp'] or environment variables
    (SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASSWORD). Returns None if absent."""
    cfg = {}
    try:
        cfg = dict(st.secrets.get("smtp", {}))
    except Exception:  # noqa: BLE001 - no secrets file
        cfg = {}
    host = cfg.get("host") or os.getenv("SMTP_HOST")
    if not host:
        return None
    return {
        "host": host,
        "port": int(cfg.get("port") or os.getenv("SMTP_PORT", 587)),
        "user": cfg.get("user") or os.getenv("SMTP_USER"),
        "password": cfg.get("password") or os.getenv("SMTP_PASSWORD"),
    }


def send_contact_email(sender: str, message: str) -> bool:
    cfg = _smtp_config()
    if cfg is None:
        return False
    msg = EmailMessage()
    msg["Subject"] = f"[PPA & DASS Settlements] Question from {sender}"
    msg["From"] = cfg["user"] or sender
    msg["To"] = CONTACT_TO
    msg["Reply-To"] = sender
    msg.set_content(
        f"Question submitted from the PPA & DASS Settlements app\n"
        f"From: {sender}\n\n{message}"
    )
    with smtplib.SMTP(cfg["host"], cfg["port"], timeout=15) as server:
        server.starttls()
        if cfg["user"] and cfg["password"]:
            server.login(cfg["user"], cfg["password"])
        server.send_message(msg)
    return True


section("Questions about these settlements?", ORANGE,
        "Leave your e-mail and your question — it goes straight to the Nexwell Power "
        "origination desk.", "✉️ We reply within 1 business day")

with st.container(border=True):
    with st.form("contact_form", clear_on_submit=False):
        f1, f2 = st.columns([1, 2])
        with f1:
            user_email = st.text_input("Your e-mail", placeholder="name@company.com")
        with f2:
            user_msg = st.text_area(
                "Your question",
                placeholder="e.g. Could you price a 10-year DASS on 20 MW starting 2027?",
                height=90)
        sent = st.form_submit_button("Send to Nexwell Power ➜", type="primary",
                                     use_container_width=True)

    if sent:
        if "@" not in user_email or "." not in user_email.split("@")[-1]:
            st.error("Please enter a valid e-mail address.")
        elif not user_msg.strip():
            st.error("Please write your question.")
        else:
            ok = False
            try:
                ok = send_contact_email(user_email.strip(), user_msg.strip())
            except Exception as exc:  # noqa: BLE001
                st.warning(f"Direct send failed ({exc}).")
            if ok:
                st.success(f"✅ Sent! Your question is on its way to {CONTACT_TO}. "
                           "We will reply to your e-mail shortly.")
            else:
                # No SMTP configured on the server → hand off to the user's mail client
                subject = urllib.parse.quote(
                    f"[PPA & DASS Settlements] Question from {user_email.strip()}")
                body = urllib.parse.quote(
                    f"From: {user_email.strip()}\n\n{user_msg.strip()}")
                st.info("Direct sending is not configured on this server — click below "
                        "to send it from your own mail client (message pre-filled).")
                st.markdown(
                    f"<a href='mailto:{CONTACT_TO}?subject={subject}&body={body}' "
                    f"target='_blank' style='display:inline-block;background:{GREEN_DARK};"
                    f"color:#fff;font-weight:700;padding:10px 22px;border-radius:10px;"
                    f"text-decoration:none;'>📨 Open e-mail to {CONTACT_TO}</a>",
                    unsafe_allow_html=True)


# ---------------------------------------------------------------------------
# Export / PDF
# ---------------------------------------------------------------------------
st.markdown("---")
st.markdown(
    """
    <div class="nx-print-card">
        <div>
            <div class="nx-print-title">Export page to PDF</div>
            <div class="nx-print-subtitle">
                Print or save the current settlement view as PDF. Recommended format:
                <b>A4 landscape</b>, margins set to <b>minimum</b>, background graphics enabled.
            </div>
        </div>
        <div class="nx-print-seal">NEXWELL POWER · SETTLEMENT VIEW</div>
    </div>
    """,
    unsafe_allow_html=True,
)
components.html(
    """
    <button onclick="window.parent.print()" style="
        width: 100%;
        padding: 14px 18px;
        border: 0;
        border-radius: 12px;
        background: #0f6b47;
        color: white;
        font-weight: 800;
        font-size: 15px;
        cursor: pointer;
        margin-top: 10px;
    ">
        🖨️ Print / Save current page as PDF
    </button>
    """,
    height=70,
)
